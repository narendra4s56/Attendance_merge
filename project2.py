from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# Load the workbook from an existing file
output_file = "./AttendanceReport.xlsx"
wb = load_workbook(output_file)

# Ensure the workbook is loaded and check its sheets
if wb is not None:
    print(f"Workbook loaded successfully. Sheets: {wb.sheetnames}")

    # Access the active sheet or explicitly access by sheet name
    ws = wb.active  # or use wb['Sheet1'] if you know the sheet name

    # Ensure that ws is not None
    if ws is not None:
        # Example of checking if a specific cell exists (has a value)
        cell_to_check = 'A1'  # Change this to any cell you want to check
        if ws[cell_to_check].value is not None:
            print(f"The value in {cell_to_check} is {ws[cell_to_check].value}")
            # Perform operations here if the cell exists
        else:
            print(f"The cell {cell_to_check} is empty.")
        
        # Apply vertical text alignment by rotating the text (90 degrees)
        cells_to_align = ['C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6', 'J6', 'K6', 'L6', 'M6', 'N6', 'O6', 'P6', 'A6', 'B6', 'T6', 'U6', 'V6', 'W6', 'X6', 'Y6', 'Z6', 'Q6', 'R6', 'S6']
        for cell in cells_to_align:
            if ws[cell].value is not None:  # Check if the cell has a value
                ws[cell].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        # Shift values only if the source cell has a value
        if ws['X5'].value is not None:
            ws['X6'] = ws['X5'].value
            ws['X5'] = ""  # Clear the original cell
            ws['X6'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        if ws['Y5'].value is not None:
            ws['Y6'] = ws['Y5'].value
            ws['Y5'] = ""
            ws['Y6'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        if ws['Z5'].value is not None:
            ws['Z6'] = ws['Z5'].value
            ws['Z5'] = ""
            ws['Z6'].alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        # Define fills for the background colors
        theory_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
        lab_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Light blue
        total_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")  # Light black (dark gray)

        # Identify and color the relevant sub-header cells
        sub_headers = {
            'Theory Percentage': theory_fill,
            'Lab Percentage': lab_fill,
            'Total Classes': total_fill,
            'Total Attended': total_fill,
            'Total Percentage': total_fill,
        }

        for row in range(2, ws.max_row + 1):  # Assuming sub-headers start from row 2
            for col in ws.iter_cols(min_row=row, max_row=row):  # Iterate through each column in the current row
                for cell in col:
                    if cell.value in sub_headers:
                        cell.fill = sub_headers[cell.value]  # Apply the corresponding fill color

        print("Cell updated, shifted, vertical text alignment applied, and colors added successfully!")

        # Save the changes to the file
        wb.save(output_file)
    else:
        print("No active worksheet found.")
else:
    print("Failed to load the workbook.")
