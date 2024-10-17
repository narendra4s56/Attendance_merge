import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


path = "./*.xlsx" 
all_files = glob.glob(path)

df_list = []


for file in all_files:
    try:
        # Load each Excel file without header to inspect the first few rows
        df = pd.read_excel(file, header=None)

        print(f"Processing file: {file}")
        print(f"Shape: {df.shape}")
        print(df.head(5))  

        if df.shape[0] < 2 or df.shape[1] < 4:  
            print(f"Skipping empty or improperly formatted file: {file}")
            continue

      
        subject_name = df.iloc[0, 0]  

        first_digit_match = re.search(r'\d', subject_name)
        # Extract the result if a match is found
        if first_digit_match:
          first_digit = int(first_digit_match.group())
          print("First single digit:", first_digit)
        else:
          print("No digit found.")

        # Set the second row as the header
        df.columns = df.iloc[1]  
        df = df.drop(index=[0, 1]) 

        # Reset index
        df.reset_index(drop=True, inplace=True)

        
        df.columns = df.columns.str.strip()  # Remove any leading/trailing spaces

        if 'Enrollment No.' not in df.columns or 'Name' not in df.columns:
            print(f"Skipping file due to missing 'Enrollment No.' or 'Name' columns: {file}")
            continue

        # Ensure numeric columns are converted to float, handling errors
        df['Total Theory'] = pd.to_numeric(df['Total Theory'], errors='coerce')
        df['Attended'] = pd.to_numeric(df['Attended'], errors='coerce')

  
        if 'Lab' in df.columns and 'Lab Attended' in df.columns:
            # Select relevant columns for both theory and lab
            df['Lab'] = pd.to_numeric(df['Lab'], errors='coerce')
            df['Lab Attended'] = pd.to_numeric(df['Lab Attended'], errors='coerce')
            df_theory = df[['Enrollment No.', 'Name', 'Total Theory', 'Attended']].copy()
            df_lab = df[['Enrollment No.', 'Name', 'Lab', 'Lab Attended']].copy()
        else:
         
            df_theory = df[['Enrollment No.', 'Name', 'Total Theory', 'Attended']].copy()
            df_lab = None  

        
        df_theory['Theory Percentage'] = (df_theory['Attended'] * 100 / df_theory['Total Theory']).fillna(0).round(2)

        # Set index for theory DataFrame
        df_theory.set_index(['Enrollment No.', 'Name'], inplace=True)
        df_theory.columns = pd.MultiIndex.from_product([[subject_name], df_theory.columns])

       
        df_list.append(df_theory)

        # If lab data exists, process it similarly
        if df_lab is not None:
            df_lab['Lab Percentage'] = (df_lab['Lab Attended'] * 100 / df_lab['Lab']).fillna(0).round(2)
            df_lab.set_index(['Enrollment No.', 'Name'], inplace=True)
            df_lab.columns = pd.MultiIndex.from_product([[subject_name], df_lab.columns])
            df_list.append(df_lab)

        # Print the DataFrames to ensure they have the correct data
        print(f"DataFrame for {subject_name} (Theory):")
        print(df_theory.head())  

        if df_lab is not None:
            print(f"DataFrame for {subject_name} (Lab):")
            print(df_lab.head())  

    except Exception as e:
        print(f"Error processing file {file}: {e}")

# Concatenate all DataFrames in the list into a single DataFrame
if df_list:
    merged_df = pd.concat(df_list, axis=1)  

   
    merged_df['Total Classes'] = merged_df.xs('Total Theory', axis=1, level=1).sum(axis=1, skipna=True) + \
                                  merged_df.xs('Lab', axis=1, level=1).sum(axis=1, skipna=True)
    merged_df['Total Attended'] = merged_df.xs('Attended', axis=1, level=1).sum(axis=1, skipna=True) + \
                                   merged_df.xs('Lab Attended', axis=1, level=1).sum(axis=1, skipna=True)

    
    total_percentage = (merged_df['Total Attended'] * 100 / merged_df['Total Classes']).replace([float('inf'), -float('inf')], 0).round(0).fillna(0)

    # Add the Total Percentage column to the DataFrame
    merged_df['Total Percentage'] = total_percentage

    # Print the merged DataFrame for verification
    print("Merged DataFrame with Total Columns and Percentages:")
    print(merged_df.head())  

    # Write the final DataFrame to a new Excel file
    output_file = "AttendanceReport.xlsx"
    merged_df.to_excel(output_file, sheet_name='Summary', index=True)  # Include index

    # Load the workbook and worksheet to add a heading
    wb = load_workbook(output_file)
    ws = wb['Summary']

    # Insert the 3 heading rows at the top
    ws.insert_rows(1, amount=4) 

    # Add text to the inserted rows
    ws['A2'] = "DEPARTMENT OF COMPUTER ENGINEERING"  
    ws['A3'] = "SESSION : JULY-DEC 2024; Semester 'A'"  
    ws['A4'] = " BTech. IYEAR ATTENDANCE SHEET"  

    # Merge cells across the columns for each heading row (adjust based on your column count)
    max_column = ws.max_column
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)  
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_column)  
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_column)  

    # Style the headings: font size, bold, and center alignment
    for row in range(1, 6):
        ws[f'A{row}'].font = Font(size=14 if row == 1 else 12, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Adjust row heights for the heading (optional)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    # Style the rest of the content as before
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Arial', size=10)


        
    # Adjust column widths for readability (if required)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30

    # Save the workbook with the new heading
    wb.save(output_file)
    print(f"Attendance report saved to {output_file} with headings")

else:
    print("No valid files to merge. No output file was created.")
